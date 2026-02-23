"""
ExcelLoader v6

Modos soportados:
  - build_from_template()  → crea Excel nuevo desde template
  - update_existing()      → actualiza Excel ya existente

Protecciones:
  - No pisa fórmulas
  - No sobreescribe valores existentes con 0
  - No escribe si value es None

Fix crítico:
  - Fuerza recálculo en Excel al abrir (fullCalcOnLoad + calcMode=auto)
    para evitar que openpyxl deje en blanco/0 los cached results de fórmulas.
"""

import shutil
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from excel.excel_mapper import ExcelMapper


class ExcelLoader:

    SHEET_931 = "931"
    SHEET_ANALISIS = "Analisis General"
    HEADER_ROW = 7
    FIRST_DATA_ROW = 8
    LAST_DATA_ROW = 38

    def __init__(self, template_path: str, consolidated_json: dict):
        self.template_path = str(Path(template_path).resolve()) if template_path else None
        self.data = consolidated_json
        self.mapper = ExcelMapper(consolidated_json)

        if "periodo_iso" not in self.data:
            raise ValueError("El JSON consolidado no contiene 'periodo_iso'.")

        self.periodo = self.data["periodo_iso"]
        self.year, self.month = [int(x) for x in self.periodo.split("-")]

    # =========================================================
    # PUBLIC METHODS
    # =========================================================

    def build_from_template(self, output_path: str):
        """
        Crea un Excel nuevo copiando el template.
        """
        if not self.template_path:
            raise ValueError("Template path no definido.")

        output_path = str(Path(output_path).resolve())

        if output_path == self.template_path:
            raise ValueError("El output no puede ser el mismo que el template.")

        shutil.copy2(self.template_path, output_path)
        print(f"Template copiado → {output_path}")

        self._process_workbook(output_path)

    def update_existing(self, excel_path: str):
        """
        Actualiza un Excel existente.
        """
        excel_path = str(Path(excel_path).resolve())

        if not Path(excel_path).exists():
            raise FileNotFoundError(f"No existe el Excel: {excel_path}")

        print(f"UPDATE → {excel_path}")
        self._process_workbook(excel_path)

    # =========================================================
    # CORE WORKBOOK PROCESSOR
    # =========================================================

    def _process_workbook(self, excel_path: str):
        # Abrimos workbook target
        wb = load_workbook(excel_path)

        # Detectar columna del mes (preferimos el workbook target, no el template)
        ag_col = self._detect_analisis_column(month=self.month, wb=wb)
        print(f"Periodo {self.periodo} → Columna Analisis General: {ag_col}")

        written_931 = self._write_931(wb)
        written_ag = self._write_analisis_general(wb, ag_col)

        total = written_931 + written_ag

        print(
            f"\nResumen:\n"
            f"  931: {written_931} celdas\n"
            f"  Analisis General: {written_ag} celdas\n"
            f"  Total: {total}"
        )

        # FIX: Forzar que Excel recalcule todo al abrir
        # (evita que queden en 0/vacío los resultados cacheados de fórmulas)
        try:
            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
        except Exception:
            # No rompe si alguna versión de openpyxl cambia
            pass

        try:
            wb.save(excel_path)
            print(f"Guardado OK → {excel_path}")
        except PermissionError:
            raise PermissionError("No se pudo guardar. ¿Está abierto en Excel?")

        wb.close()

    # =========================================================
    # WRITE 931
    # =========================================================

    def _write_931(self, wb) -> int:
        if self.SHEET_931 not in wb.sheetnames:
            raise ValueError(f"No existe la hoja '{self.SHEET_931}'")

        ws = wb[self.SHEET_931]
        values = self.mapper.get_931_values(self.month)

        print(f"\nEscribiendo hoja '{self.SHEET_931}'")

        written = 0

        for item in values:
            row = item["row"]
            col = item["col"]
            value = item["value"]
            label = item["label"]

            cell = ws.cell(row=row, column=col)

            # No pisar fórmulas
            if cell.data_type == "f":
                continue

            # No sobreescribir con 0 si ya hay dato
            if isinstance(value, (int, float)) and value == 0:
                if cell.value not in (None, "", 0):
                    continue

            # No escribir None
            if value is None:
                continue

            cell.value = value

            col_letter = get_column_letter(col)
            tipo = "texto" if isinstance(value, str) else "numérico"
            print(f"  {col_letter}{row} | {label} → {value} ({tipo})")
            written += 1

        return written

    # =========================================================
    # WRITE ANALISIS GENERAL
    # =========================================================

    def _write_analisis_general(self, wb, month_col: int) -> int:
        if self.SHEET_ANALISIS not in wb.sheetnames:
            raise ValueError(f"No existe la hoja '{self.SHEET_ANALISIS}'")

        ws = wb[self.SHEET_ANALISIS]

        print(f"\nEscribiendo hoja '{self.SHEET_ANALISIS}' col {month_col}")

        written = 0

        for row_num in range(self.FIRST_DATA_ROW, self.LAST_DATA_ROW + 1):
            concepto_cell = ws.cell(row=row_num, column=2)  # B
            if not concepto_cell.value:
                continue

            concepto_text = str(concepto_cell.value).strip()
            target_cell = ws.cell(row=row_num, column=month_col)

            # No pisar fórmulas
            if target_cell.data_type == "f":
                continue

            value = self.mapper.resolve_analisis_value(row_num, concepto_text)

            if value is None:
                continue

            # No sobreescribir con 0 si ya hay dato
            if isinstance(value, (int, float)) and value == 0:
                if target_cell.value not in (None, "", 0):
                    continue

            target_cell.value = value

            tipo = "texto" if isinstance(value, str) else "numérico"
            print(f"  Fila {row_num} | {concepto_text} → {value} ({tipo})")
            written += 1

        return written

    # =========================================================
    # DETECT COLUMN
    # =========================================================

    def _detect_analisis_column(self, month: int, wb=None) -> int:
        """
        Prioridad:
          1) Buscar en workbook target (wb) por datetime en fila HEADER_ROW
          2) Fallback a template (Títulos!B3 -> month+3)
        """

        # ---- 1) Buscar en WB target (más confiable si estás actualizando un Excel ya armado)
        if wb is not None and self.SHEET_ANALISIS in wb.sheetnames:
            ws = wb[self.SHEET_ANALISIS]
            for col_cell in ws[self.HEADER_ROW]:
                val = col_cell.value
                if isinstance(val, datetime):
                    if val.year == self.year and val.month == month:
                        return col_cell.column

        # ---- 2) Fallback a template por ancla Títulos!B3
        if not self.template_path:
            raise ValueError("No se pudo detectar columna: no hay wb válido ni template_path.")

        wb_tpl = load_workbook(self.template_path, data_only=False)
        if "Títulos" not in wb_tpl.sheetnames:
            wb_tpl.close()
            raise ValueError("No existe hoja 'Títulos' en el template.")

        anchor = wb_tpl["Títulos"].cell(row=3, column=2).value  # B3
        wb_tpl.close()

        if isinstance(anchor, datetime):
            col = month + 3  # D=4 para enero, ... O=15 para dic
            if 4 <= col <= 15:
                return col

        raise ValueError(
            f"No se pudo determinar la columna para mes {month}. "
            "Verificar headers en Analisis General o ancla Títulos!B3."
        )